from sqlescapy import sqlescape
from openpyxl import load_workbook
from SZChat_funcoes import fgetCelular

def main():
    vNomeArquivoExcel = 'C:\docker\code\excel\cliente_sem_telefonia.xlsx'
    print(vNomeArquivoExcel)

    # importar excel
    arquivo_excel = load_workbook(vNomeArquivoExcel)
    planilha1 = arquivo_excel.active
    
    max_linha = planilha1.max_row
    max_coluna = planilha1.max_column
    
    for i in range(1, max_linha + 1):
        #print(planilha1.cell(row=i, column=4)) 
        #print( planilha1.cell(row=i, column=3))
        
        vFone = str(planilha1.cell(row=i, column=3).value)
        vCelu = str(planilha1.cell(row=i, column=4).value)
        vFoneFormatado = fgetCelular( vFone, vCelu )
        planilha1.cell(row=i, column=5, value=vFoneFormatado) 
        print('FONE: ' + vFoneFormatado)

    arquivo_excel.save(vNomeArquivoExcel)
    arquivo_excel.close()

if __name__ == '__main__':
  main()
  