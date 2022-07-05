import sys
import logging
import random
from time import sleep
import SZChat_funcoes
import SynSuite_funcoes
import bd_conecta
#import funcoes
import json
import pyparsing
import time
import hora
from sqlescapy import sqlescape
from openpyxl import load_workbook

def getDadosPessoa(vID: str, vBD):
    sql = f"""select 
                p.name ,
                p.phone ,
                p.cell_phone_1 ,
                p.cell_phone_2 ,
                p.cell_phone_1_has_whatsapp ,
                p.cell_phone_2_has_whatsapp ,
                p.fax_phone ,
                p.commercial_phone ,
                p.whatsapp_phone 
            from people p 
                inner join contracts c 
                    on c.client_id = p.id  
            where c.id = {vID}
    """
    try:
        cursor = vBD.cursor()
        cursor.execute(sql)
        rs = cursor.fetchall()
        cursor.close()
        return rs
    except Exception as e:
        logging.error('ONUS.PY' + str(e))
        return     

def main():
    vNomeArquivoExcel = 'C:\docker\code\ONUSPADRAO.xlsx'
    print(vNomeArquivoExcel)

    # importar excel
    arquivo_excel = load_workbook(vNomeArquivoExcel)
    planilha1 = arquivo_excel.active
    
    max_linha = planilha1.max_row
    max_coluna = planilha1.max_column
    
    bd = bd_conecta.conecta_db_tche()
    
    for i in range(1, max_linha + 1):
        vCliente = planilha1.cell(row=i, column=3).value
        retorno = getDadosPessoa(vCliente, bd)
        if retorno != None:
            for r in retorno:
                planilha1.cell(row=i, column=6, value=r['name']) 
                planilha1.cell(row=i, column=7, value=r['phone']) 
                planilha1.cell(row=i, column=8, value=r['cell_phone_1']) 
                planilha1.cell(row=i, column=9, value=r['cell_phone_2']) 
                planilha1.cell(row=i, column=10, value=r['cell_phone_1_has_whatsapp']) 
                planilha1.cell(row=i, column=11, value=r['cell_phone_2_has_whatsapp']) 
                planilha1.cell(row=i, column=12, value=r['fax_phone']) 
                planilha1.cell(row=i, column=13, value=r['commercial_phone']) 
                planilha1.cell(row=i, column=14, value=r['whatsapp_phone']) 
            print(vCliente)

    arquivo_excel.save(vNomeArquivoExcel)
    arquivo_excel.close()

    bd.close()

if __name__ == '__main__':
  main()
  