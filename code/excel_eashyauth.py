import sys
import logging
import random
from time import sleep
import SZChat_funcoes
import SynSuite_funcoes
import bd_conecta
#import funcoes
import json
import pyparsing
import time
import hora
from sqlescapy import sqlescape
from openpyxl import load_workbook

def getDadosPessoa(vFONE: str, vBD):
    sql = f"""select 
	p.id ,
	p.name,
	REPLACE(
		REPLACE(
			REPLACE(	
				REPLACE(cast(coalesce(p.phone,'') as varchar(50)), '(',''),
				')',''
				   ),
				'-',''),
			' ','') as PHONE,
	REPLACE(
		REPLACE(
			REPLACE(	
				REPLACE(cast(coalesce(p.cell_phone_1,'') as varchar(50)), '(',''),
				')',''
				   ),
				'-',''),
			' ','') as CELL_PHONE
from people p 
  inner join contracts c 
    on c.client_id = p.id 
where c.v_status <> 'Cancelado'
   and REPLACE(
		REPLACE(
			REPLACE(	
				REPLACE(cast(coalesce(p.phone,'') as varchar(50)), '(',''),
				')',''
				   ),
				'-',''),
			' ','') = '{vFONE}'
	OR			
	REPLACE(
		REPLACE(
			REPLACE(	
				REPLACE(cast(coalesce(p.cell_phone_1,'') as varchar(50)), '(',''),
				')',''
				   ),
				'-',''),
			' ','') = '{vFONE}'			
  
    """
    try:
        cursor = vBD.cursor()
        cursor.execute(sql)
        rs = cursor.fetchall()
        cursor.close()
        return rs
    except Exception as e:
        print(str(e))
        return     

def main():
    vNomeArquivoExcel = 'C:\docker\code\easyauth.xlsx'
    print(vNomeArquivoExcel)

    # importar excel
    arquivo_excel = load_workbook(vNomeArquivoExcel)
    planilha1 = arquivo_excel.active
    
    max_linha = planilha1.max_row
    max_coluna = planilha1.max_column
    
    bd = bd_conecta.conecta_db_tche()
    
    for i in range(1, max_linha + 1):
        vFONE = planilha1.cell(row=i, column=4).value
        retorno = getDadosPessoa(vFONE, bd)
        print(retorno)
        if len(retorno) == 0:
            planilha1.cell(row=i, column=7, value='NAO') 
        else:
            planilha1.cell(row=i, column=7, value='SIM') 
            
        print(vFONE)
    bd.close()
    arquivo_excel.save(vNomeArquivoExcel)
    arquivo_excel.close()

if __name__ == '__main__':
  main()
  