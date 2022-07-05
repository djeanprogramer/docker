import sys
import logging
import random
from time import sleep
import SZChat_funcoes
import SynSuite_funcoes
import bd_conecta
#import funcoes
import json
import pyparsing
import time
import hora
from sqlescapy import sqlescape
from openpyxl import load_workbook

def getDadosPessoa(vNome: str, vBD):
    sql = f"""select 
                c.client_id  
            from people p 
            inner join contracts c 
                on c.client_id = p.id
            where upper(p.name) = upper('{vNome}')
              and c.v_status <> 'Cancelado'
            limit 1
    """
    try:
        cursor = vBD.cursor()
        cursor.execute(sql)
        rs = cursor.fetchall()
        cursor.close()
        return rs
    except Exception as e:
        logging.error('CLARO.PY' + str(e))
        return     

def main():
    vNomeArquivoExcel = 'C:\docker\excel\claro.xlsx'
    print(vNomeArquivoExcel)

    # importar excel
    arquivo_excel = load_workbook(vNomeArquivoExcel)
    planilha1 = arquivo_excel.active
    print('Load Excel')
    
    max_linha = planilha1.max_row
    max_coluna = planilha1.max_column
    
    bd = bd_conecta.conecta_db_tche()
    print('Conectou BD')
    
    for i in range(1, max_linha + 1):
        vCliente = planilha1.cell(row=i, column=8).value
        retorno = getDadosPessoa(vCliente, bd)
        if retorno != None:
            for r in retorno:
                planilha1.cell(row=i, column=1, value=r['client_id']) 
                print(vCliente + ' - SIM')
        else:
            planilha1.cell(row=i, column=1, value='0') 
            print(vCliente + ' - NÃO')

    arquivo_excel.save(vNomeArquivoExcel)
    arquivo_excel.close()
    print('Fechou Excel')

    bd.close()

if __name__ == '__main__':
  main()
  